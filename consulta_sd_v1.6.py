#%%
import pandas as pd
import openpyxl
import os
from openpyxl.styles import Alignment, Font
from openpyxl.cell import Cell

import streamlit as st


st.title("📄 Reporte Solicitudes Neomante")
uploaded_file = st.file_uploader("Sube tu reporte de Neomante (.xlsx)", type=("xlsx"))


libro_sd=openpyxl.load_workbook(uploaded_file)
nom_hojas=libro_sd.sheetnames
df_total=pd.DataFrame()
df_aux=pd.DataFrame()



df_ift_se=pd.read_excel("reporte_se_ubicacion.xlsx")
df_ift_lin=pd.read_excel("reporte_lineas_ubicacion.xlsx")
df_ift_gen=pd.read_excel("reporte_centrales_ubicacion.xlsx")


dic_remp={"<p>":"","</p>":"\n","<li>":"\n","</li>":"","<ul>":"","</ul>":"","<strong>":"","</strong>":"", "&nbsp;":" ","<i>":"","</i>":""}


for i in nom_hojas:
    df=pd.read_excel(path+"temp_sd.xlsx", sheet_name=i)
    nom=df.loc[0][0][26:].replace(":","_")
    columns=df.loc[5]
    columns=columns.reset_index(drop=True)
    df=df.rename(columns=dict(zip(df.columns, columns)))
    df=df.drop(df.index[:6])
    df=df.reset_index(drop=True)
    
    if i=="Subestacion":
        df.columns.values[8] = 'id'
        df=pd.merge(df,df_ift_se, on="id", how="left")
        #df.columns.values[-1] = 'Reg'
        df_aux_se=df
        df=df[["Número","Empresa","Tipo Solicitud","Tipo Programación","SubEstación","Elemento(s)","region_nombre","Trabajos a Realizar","Descripción Nivel Riesgo","Comentario Adicional","Consumo","Fecha Inicio","Fecha Fin","Fecha Efectiva Inicio","Fecha Efectiva Fin","Comentarios DAOP","Estado DAOP","Historial de Estados","Empresas Afectadas"]]
        df=df.rename(columns={"SubEstación":"Elemento","Elemento(s)":"Descripción"})

    if i=="Central Generadora":
        df.columns.values[8] = 'id'
        df=pd.merge(df,df_ift_gen, on="id", how="left")
        #df.columns.values[-1] = 'Reg'
        df_aux_central=df
        df=df[["Número","Empresa","Tipo Solicitud","Tipo Programación","Central","Unidad(es)","region_nombre","Trabajos a Realizar","Descripción Nivel Riesgo","Comentario Adicional","Consumo","Fecha Inicio","Fecha Fin","Fecha Efectiva Inicio","Fecha Efectiva Fin","Comentarios DAOP","Estado DAOP","Historial de Estados","Empresas Afectadas"]]
        df=df.rename(columns={"Central":"Elemento","Unidad(es)":"Descripción"})
    if i=="Linea":
        df.columns.values[8] = 'id'
        df=pd.merge(df,df_ift_lin, on="id", how="left")
        #df.columns.values[-1] = 'Reg'
        df_aux_linea=df
        df=df[["Número","Empresa","Tipo Solicitud","Tipo Programación","Línea","Tramo(s)","region_nombre","Trabajos a Realizar","Descripción Nivel Riesgo","Comentario Adicional","Consumo","Fecha Inicio","Fecha Fin","Fecha Efectiva Inicio","Fecha Efectiva Fin","Comentarios DAOP","Estado DAOP","Historial de Estados","Empresas Afectadas"]]
        df=df.rename(columns={"Línea":"Elemento","Tramo(s)":"Descripción"})

    df_total=pd.concat([df_total,df],axis=0)

#reemplazar quitar el caracter "DI" en todas las filas de la columna "Número"
df_total["Número"]=df_total["Número"].str.replace("DI","", regex=False) 

df_total=df_total[df_total["Tipo Programación"]!="Curso Forzoso"]

df_total=df_total.drop(columns=["Tipo Programación"], errors='ignore')

df_total=df_total.rename(columns={"region_nombre":"Región"})

df_total=df_total.fillna("")
df_total["Comentarios DAOP"]=df_total["Comentarios DAOP"].replace(dic_remp, regex=True)

df_estados=df_total["Historial de Estados"].str.split("Pendiente", expand=True)
df_estados=df_estados[1].str.split(",", expand=True)
df_total["Historial de Estados"]=df_estados[1].str.replace("•Fecha: ","")
df_total["Historial de Estados"]=pd.to_datetime(df_total["Historial de Estados"],dayfirst=True)
df_total=df_total.rename(columns={"Historial de Estados":"Fecha Envío"})


df_total["Número"]='=HYPERLINK("https://neomante.coordinador.cl/desconexion_intervencion/lista?correlativo='+df_total["Número"].astype(str)+'","'+df_total["Número"].astype(str)+'")'
df_total["Horas"]=pd.to_datetime(df_total["Fecha Fin"])-pd.to_datetime(df_total["Fecha Inicio"])
df_total["Horas"]=df_total["Horas"].dt.total_seconds()/3600




os.remove(path+"temp_sd.xlsx")

df_total.to_excel("salidas_sd_"+nom+".xlsx", index=False)

libro = openpyxl.load_workbook(r"salidas_sd_"+nom+".xlsx")
sheet = libro.active 

for column in ("A", "B", "C", "D", "E", "F", "P", "Q"): sheet.column_dimensions[column].width = 18
for column in ("G", "H", "I", "O"): sheet.column_dimensions[column].width = 32
for column in ("K", "L","Q"): sheet.column_dimensions[column].width = 22


for row in sheet.iter_rows():  
    for cell in row:      
        cell.alignment = Alignment(wrap_text=True,vertical='top') 


for cell in sheet["C"]:
    if cell.value == "Intervención":
        cell.font=Font(color="AD8000", bold=True)
    if cell.value == "Desconexión":
        cell.font=Font(color="FF0000", bold=True)        
        


for cell in sheet["F"]:
    if (cell.value == "Antofagasta")|(cell.value == "Arica y Parinacota")|(cell.value == "Tarapacá"):
        cell.font=Font(color="0070C0", bold=True)
    
    if (cell.value == "Atacama")|(cell.value == "Coquimbo")|(cell.value == "Valparaíso"):
        cell.font=Font(color="C00000", bold=True)
    



sheet["F1"].font=Font(color="000000", bold=True)


sheet=libro.create_sheet("Consulta Correo")

sheet["A1"]="Número"
sheet["B1"]="Número"
sheet["C1"]="Empresa"
sheet["D1"]="Elemento"
sheet["E1"]="Trabajos a Realizar"
sheet["F1"]="Descripción Nivel Riesgo"
sheet["G1"]="Comentario Adicional"
sheet["H1"]="Fecha Inicio"
sheet["I1"]="Fecha Fin"

sheet["B2"]="=A2"
sheet["C2"]="=VLOOKUP($A2,Sheet1!$A$2:$T$"+str(len(df_total))+",2,FALSE)"
sheet["D2"]="=VLOOKUP($A2,Sheet1!$A$2:$T$"+str(len(df_total))+",5,FALSE)"
sheet["E2"]="=VLOOKUP($A2,Sheet1!$A$2:$T$"+str(len(df_total))+",7,FALSE)"
sheet["F2"]="=VLOOKUP($A2,Sheet1!$A$2:$T$"+str(len(df_total))+",8,FALSE)"
sheet["G2"]="=VLOOKUP($A2,Sheet1!$A$2:$T$"+str(len(df_total))+",9,FALSE)"
sheet["H2"]="=VLOOKUP($A2,Sheet1!$A$2:$T$"+str(len(df_total))+",11,FALSE)"
sheet["I2"]="=VLOOKUP($A2,Sheet1!$A$2:$T$"+str(len(df_total))+",12,FALSE)"




for column in ("A", "B", "C", "D"): sheet.column_dimensions[column].width = 18
for column in ("E", "F", "G"): sheet.column_dimensions[column].width = 32
for column in ("H","I"): sheet.column_dimensions[column].width = 22


for row in sheet.iter_rows():  
    for cell in row:      
        cell.alignment = Alignment(wrap_text=True,vertical='top') 




libro.save(r"salidas_sd_"+nom+".xlsx")


archivo = "salidas_sd_"+nom+".xlsx"
wb = load_workbook(archivo)
ws = wb["Sheet1"]  # Puedes cambiarlo por ws = wb['NombreHoja'] si conoces el nombre

# Definir la lista de valores permitidos
opciones = ["Aprobado", "Rechazado", "Pendiente", "Aprobado sin activación"]
lista = DataValidation(
    type="list",
    formula1='"{}"'.format(",".join(opciones)),
    allow_blank=True,  # Permite dejar la celda vacía si es necesario
    showDropDown=True
)

# Agregar la validación a la hoja
ws.add_data_validation(lista)

# Recorrer las celdas vacías de la columna B (por ejemplo, de la fila 2 a 100)
for row in range(2, 500):
    celda = ws[f'P{row}']
    if celda.value is None or str(celda.value).strip() == "":
        lista.add(celda)

# Guardar el archivo

ws.freeze_panes = "A2"

max_col = ws.max_column
col_letter = ws.cell(row=1, column=max_col).column_letter
ws.auto_filter.ref = f"A1:{col_letter}1"


for fila in range(2, ws.max_row + 1):
    ws.row_dimensions[fila].height = 100



wb.save("salidas_sd_"+nom+".xlsx")

