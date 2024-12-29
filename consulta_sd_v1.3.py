#%%
import pandas as pd
import openpyxl
import os
from openpyxl.styles import Alignment, Font
from openpyxl.cell import Cell

import streamlit as st


st.title("📄 Reporte Solicitudes Neomante")
st.write(
    "Subtitulo"
)

uploaded_file = st.file_uploader("Sube tu reporte de Neomante (.xlsx)", type=("xlsx"))


libro_sd=openpyxl.load_workbook(uploaded_file)
nom_hojas=libro_sd.sheetnames
df_total=pd.DataFrame()
df_aux=pd.DataFrame()
df_ift_se=pd.read_excel("reporte_subestaciones.xlsx", sheet_name="Subestaciones")
df_ift_se=df_ift_se[["Nombre", "Región","Comuna"]]
df_ift_se=df_ift_se[["Nombre", "Región","Comuna"]]
df_ift_se=df_ift_se.rename(columns={"Nombre":"SubEstación"})

df_ift_lin=pd.read_excel("reporte_lineas.xlsx", sheet_name="Lineas")
df2=df_ift_lin["Nombre"]
df2=df2.str.replace(" - ", "_")
df2=df2.str.replace(" – ", "_")
df_ift_lin=df_ift_lin.join((df2.str.split("_", expand=True))[0])
df_ift_lin[0]="S/E "+df_ift_lin[0]
df_ift_lin=df_ift_lin.rename(columns={0:"SubEstación", "Nombre":"Línea"})
df_ift_lin=df_ift_lin[["Línea", "SubEstación"]]
df_ift_lin=pd.merge(df_ift_lin,df_ift_se, on="SubEstación")
df_ift_lin=df_ift_lin.drop_duplicates()


df_ift_gen=pd.read_excel("reporte_generadores.xlsx", sheet_name="Generadores")
df_ift_gen=df_ift_gen[["Nombre Central", "Subestación de inyección"]]
df_ift_gen=df_ift_gen.rename(columns={"Subestación de inyección":"SubEstación", "Nombre Central":"Central"})
df_ift_gen=pd.merge(df_ift_gen,df_ift_se, on="SubEstación")
df_ift_gen=df_ift_gen.drop_duplicates()


dic_remp={"<p>":"","</p>":"\n","<li>":"\n","</li>":"","<ul>":"","</ul>":"","<strong>":"","</strong>":"", "&nbsp;":" ","<i>":"","</i>":""}


for i in nom_hojas:
    df=pd.read_excel(uploaded_file, sheet_name=i)
    nom=df.loc[0][0][26:].replace(":","_")
    columns=df.loc[5]
    columns=columns.reset_index(drop=True)
    df=df.rename(columns=dict(zip(df.columns, columns)))
    df=df.drop(df.index[:6])
    df=df.reset_index(drop=True)
    
    if i=="Subestacion":
        df=pd.merge(df,df_ift_se, on="SubEstación", how="left")
        df=df[["Número","Empresa","Tipo Solicitud","Tipo Programación","SubEstación","Elemento(s)","Región","Trabajos a Realizar","Descripción Nivel Riesgo","Comentario Adicional","Consumo","Fecha Inicio","Fecha Fin","Fecha Efectiva Inicio","Fecha Efectiva Fin","Comentarios DAOP","Estado DAOP","Historial de Estados","Empresas Afectadas"]]
        df=df.rename(columns={"SubEstación":"Elemento","Elemento(s)":"Descripción"})

    if i=="Central Generadora":
        df=pd.merge(df,df_ift_gen, on="Central", how="left")
        df=df[["Número","Empresa","Tipo Solicitud","Tipo Programación","Central","Unidad(es)","Región","Trabajos a Realizar","Descripción Nivel Riesgo","Comentario Adicional","Consumo","Fecha Inicio","Fecha Fin","Fecha Efectiva Inicio","Fecha Efectiva Fin","Comentarios DAOP","Estado DAOP","Historial de Estados","Empresas Afectadas"]]
        df=df.rename(columns={"Central":"Elemento","Unidad(es)":"Descripción"})
    if i=="Linea":
        df_aux["SubEstación"]=df["Línea"]
        df_aux["SubEstación"]=df_aux["SubEstación"].str.replace(" - ", "_")
        df_aux["SubEstación"]=df_aux["SubEstación"].str.replace(" – ", "_")
        df_aux["SubEstación"]=df_aux["SubEstación"].str.split("_", expand=True)[0]
        df_aux["SubEstación"]="S/E "+df_aux["SubEstación"]
        df=df.join(df_aux)
        df=pd.merge(df,df_ift_se, on="SubEstación", how="left")
        df=df[["Número","Empresa","Tipo Solicitud","Tipo Programación","Línea","Tramo(s)","Región","Trabajos a Realizar","Descripción Nivel Riesgo","Comentario Adicional","Consumo","Fecha Inicio","Fecha Fin","Fecha Efectiva Inicio","Fecha Efectiva Fin","Comentarios DAOP","Estado DAOP","Historial de Estados","Empresas Afectadas"]]
        df=df.rename(columns={"Línea":"Elemento","Tramo(s)":"Descripción"})
    df_total=pd.concat([df_total,df],axis=0)


df_total=df_total.fillna("")
df_total["Comentarios DAOP"]=df_total["Comentarios DAOP"].replace(dic_remp, regex=True)

df_estados=df_total["Historial de Estados"].str.split("Pendiente", expand=True)
df_estados=df_estados[1].str.split(",", expand=True)
df_total["Historial de Estados"]=df_estados[1].str.replace("•Fecha: ","")
df_total["Historial de Estados"]=pd.to_datetime(df_total["Historial de Estados"],dayfirst=True)
df_total=df_total.rename(columns={"Historial de Estados":"Fecha Envío"})


df_total["Número"]='=HYPERLINK("https://neomante.coordinador.cl/desconexion_intervencion/lista?correlativo='+df_total["Número"].astype(str)+'","'+df_total["Número"].astype(str)+'")'
df_total["Horas"]=pd.to_datetime(df_total["Fecha Fin"])-pd.to_datetime(df_total["Fecha Inicio"])
df_total["Horas"]=df_total["Horas"].dt.total_seconds()/3600



df_total.to_excel("salidas_sd_"+nom+".xlsx", index=False)

libro = openpyxl.load_workbook(r"salidas_sd_"+nom+".xlsx")
sheet = libro.active 

for column in ("A", "B", "C", "D", "E", "F", "P", "G", "Q"): sheet.column_dimensions[column].width = 18
for column in ("H", "I", "J", "P"): sheet.column_dimensions[column].width = 32
for column in ("L","M", "R"): sheet.column_dimensions[column].width = 22


for row in sheet.iter_rows():  
    for cell in row:      
        cell.alignment = Alignment(wrap_text=True,vertical='top') 


for cell in sheet["C"]:
    if cell.value == "Intervención":
        cell.font=Font(color="AD8000", bold=True)
    if cell.value == "Desconexión":
        cell.font=Font(color="FF0000", bold=True)        
        
for cell in sheet["Q"]:
    if cell.value == "Aprobado":
        cell.font=Font(color="66C6CC", bold=True)
    if cell.value == "Pendiente":
        cell.font=Font(color="F8A41D", bold=True)     
    if cell.value == "Rechazado":
        cell.font=Font(color="E18479", bold=True)  
    if cell.value == "Aprobado sin activación":
        cell.font=Font(color="B9DA36", bold=True) 


for cell in sheet["G"]:
    if (cell.value == "ANTOFAGASTA")|(cell.value == "ARICA Y PARINACOTA")|(cell.value == "ATACAMA")|(cell.value == "COQUIMBO")|(cell.value == "TARAPACÁ")|(cell.value == "VALPARAÍSO"):
        cell.font=Font(color="0070C0", bold=True)
    else:
        cell.font=Font(color="00B050", bold=True)



sheet["G1"].font=Font(color="000000", bold=True)


sheet=libro.create_sheet("Consulta Correo")

sheet["A1"]="Número"
sheet["B1"]="Número"
sheet["C1"]="Empresa"
sheet["D1"]="Elemento"
sheet["E1"]="Trabajos a Realizar"
sheet["F1"]="Descripción Nivel Riesgo"
sheet["G1"]="Comentario Adicional"
sheet["H1"]="Fecha Inicio"
sheet["I1"]="Fecha Fin"

sheet["B2"]="=A2"
sheet["C2"]="=VLOOKUP($A2,Sheet1!$A$2:$T$"+str(len(df_total))+",2,FALSE)"
sheet["D2"]="=VLOOKUP($A2,Sheet1!$A$2:$T$"+str(len(df_total))+",5,FALSE)"
sheet["E2"]="=VLOOKUP($A2,Sheet1!$A$2:$T$"+str(len(df_total))+",8,FALSE)"
sheet["F2"]="=VLOOKUP($A2,Sheet1!$A$2:$T$"+str(len(df_total))+",9,FALSE)"
sheet["G2"]="=VLOOKUP($A2,Sheet1!$A$2:$T$"+str(len(df_total))+",10,FALSE)"
sheet["H2"]="=VLOOKUP($A2,Sheet1!$A$2:$T$"+str(len(df_total))+",12,FALSE)"
sheet["I2"]="=VLOOKUP($A2,Sheet1!$A$2:$T$"+str(len(df_total))+",13,FALSE)"




for column in ("A", "B", "C", "D"): sheet.column_dimensions[column].width = 18
for column in ("E", "F", "G"): sheet.column_dimensions[column].width = 32
for column in ("H","I"): sheet.column_dimensions[column].width = 22


for row in sheet.iter_rows():  
    for cell in row:      
        cell.alignment = Alignment(wrap_text=True,vertical='top') 


libro.save(r"salidas_sd_"+nom+".xlsx")



with open("salidas_sd_"+nom+".xlsx", "rb") as xlsx:
    btn = st.download_button(
        label="Descargar XLSX",
        data=xlsx,
        file_name="salidas_sd_"+nom+".xlsx",
        mime="image/png",
    
    )

os.remove("salidas_sd_"+nom+".xlsx")

# %%
